"""Adapters must produce STABLE payloads: same content -> same payload, even
when the bytes differ (re-saves, encodings); different content -> different."""
import pytest

from nightward.adapters import from_docx, from_file, from_pdf, from_text, from_xlsx
from nightward.errors import NightwardError

# --- from_file (stdlib, any format) ------------------------------------------


def test_from_file_fingerprints_any_binary(tmp_path):
    blob = tmp_path / "report.hwp"  # no parser exists; artifact gate still works
    data = b"\xd0\xcf\x11\xe0 fake hwp payload"
    blob.write_bytes(data)
    payload = from_file(blob)
    assert payload["size_bytes"] == len(data)
    assert len(payload["sha256"]) == 64
    assert from_file(blob) == payload  # deterministic


def test_from_file_missing_is_clean_error(tmp_path):
    with pytest.raises(NightwardError):
        from_file(tmp_path / "nope.bin")


# --- from_text (encoding detection) -------------------------------------------


def test_from_text_same_content_different_encodings_gate_equal(tmp_path):
    content = "한글 안내문\n둘째 줄"
    a = tmp_path / "utf8.txt"
    b = tmp_path / "cp949.txt"
    a.write_bytes(content.encode("utf-8"))
    b.write_bytes(content.encode("cp949"))
    pa, pb = from_text(a), from_text(b)
    assert (pa["encoding"], pb["encoding"]) == ("utf-8", "cp949")
    assert pa["text_sha256"] == pb["text_sha256"]  # content equality survives encoding
    assert pa["chars"] == pb["chars"]


def test_from_text_undecodable_falls_back_to_artifact(tmp_path):
    f = tmp_path / "junk.txt"
    f.write_bytes(b"\xff\xfe\xff\x00\xff")
    payload = from_text(f, encodings=("utf-8",))
    assert payload["encoding"] == "unknown"
    assert payload["size_bytes"] == 5
    assert "sha256" in payload


# --- from_pdf ------------------------------------------------------------------


def test_from_pdf_content_stable_across_byte_level_resave(tmp_path):
    pypdf = pytest.importorskip("pypdf")
    a = tmp_path / "a.pdf"
    b = tmp_path / "b.pdf"
    w = pypdf.PdfWriter()
    w.add_blank_page(width=200, height=200)
    w.add_blank_page(width=200, height=200)
    with a.open("wb") as fh:
        w.write(fh)
    # re-save with metadata churn: bytes differ, content doesn't
    w2 = pypdf.PdfWriter()
    for page in pypdf.PdfReader(str(a)).pages:
        w2.add_page(page)
    w2.add_metadata({"/Producer": "resaver-9000"})
    with b.open("wb") as fh:
        w2.write(fh)

    assert from_file(a) != from_file(b)        # artifact gate sees the noise
    pa, pb = from_pdf(a), from_pdf(b)
    assert pa == pb                            # content gate does not
    assert pa["pages"] == 2


# --- from_docx -------------------------------------------------------------------


def test_from_docx_counts_and_content_hash(tmp_path):
    docx = pytest.importorskip("docx")

    def build(path):
        d = docx.Document()
        d.add_paragraph("보고서 제목")
        d.add_paragraph("본문 문단")
        d.add_table(rows=2, cols=3)
        d.save(str(path))

    a = tmp_path / "a.docx"
    b = tmp_path / "b.docx"
    build(a)
    build(b)
    pa, pb = from_docx(a), from_docx(b)
    assert pa == pb                            # separately built, same content
    assert pa["tables"] == 1
    assert pa["text_chars"] > 0


# --- from_xlsx --------------------------------------------------------------------


def _build_xlsx(path, total=42):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "비교"
    ws["A1"], ws["B1"] = "항목", "값"
    ws["A2"], ws["B2"] = "등록금", total
    wb.save(str(path))


def test_from_xlsx_same_values_gate_equal_one_cell_breaches(tmp_path):
    pytest.importorskip("openpyxl")
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    c = tmp_path / "c.xlsx"
    _build_xlsx(a)
    _build_xlsx(b)
    _build_xlsx(c, total=43)  # one cell moved
    pa, pb, pc = from_xlsx(a), from_xlsx(b), from_xlsx(c)
    assert pa == pb                            # zip timestamps don't leak into the payload
    assert pa["sheets"] == {"비교": {"rows": 2, "cols": 2}}
    assert pa["content_sha256"] != pc["content_sha256"]
    assert pa["sheets"] == pc["sheets"]        # structure same; only values moved
